import os
from shutil import copyfile
from datetime import datetime
import mysql.connector as mariadb
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

sql = """SELECT
	grp1.first_broker, COUNT(grp1.identifier) AS new_tenderers_count
FROM
	(
		SELECT
			ts.`identifier`,
			(
				SELECT br2.code
				FROM
					tenders t2
					LEFT JOIN `brokers` br2 ON br2.`id` = t2.`broker_id`
					LEFT JOIN bids b2 ON b2.`tender_id` = t2.id
					LEFT JOIN `tenderers_bids` tb2 ON tb2.`bid_id` = b2.id
				WHERE tb2.tenderer_id = ts.`id`
				ORDER BY b2.bid_date
				LIMIT 1) AS first_broker,
				COUNT(t.id) AS tenders_count
		FROM
			tenders t
			LEFT JOIN `brokers` br ON br.`id` = t.`broker_id`
			LEFT JOIN bids b ON b.`tender_id` = t.id
			LEFT JOIN `tenderers_bids` tb ON tb.`bid_id` = b.id
			LEFT JOIN tenderers ts ON ts.`id` = tb.`tenderer_id`
		WHERE ts.`id` IS NOT NULL
		GROUP BY ts.`identifier`
		HAVING COUNT(t.id) > 1
		ORDER BY ts.`id`
	) AS grp1
GROUP BY grp1.first_broker
ORDER BY 2 desc""";

data = []

conn = mariadb.connect(host='127.0.0.1', user='root', password='root', database='reports_data', charset='utf8')
cursor = conn.cursor(buffered=True)

res = cursor.execute(sql)
for (broker_name, suppliers_count) in cursor:
  data.append([broker_name, suppliers_count])

cursor.close()
conn.close()

templates_dir = 'reports/templates'
result_dir = 'reports'
template_file_name = 'one.xlsx'

t = os.path.splitext(template_file_name)
result_file = os.path.join(result_dir, t[0] +'-'+ datetime.now().strftime('%Y-%m-%d-%H-%M-%S') + t[1])

copyfile(os.path.join(templates_dir, template_file_name), result_file)

wb = load_workbook(filename = result_file)
ws = wb.active

row = 2

for (broker_name, suppliers_count) in data:
  ws.cell(row=row, column=1, value=broker_name)
  ws.cell(row=row, column=2, value=suppliers_count)
  row += 1
  print("{} - {}".format(broker_name, suppliers_count))

wb.save(result_file)