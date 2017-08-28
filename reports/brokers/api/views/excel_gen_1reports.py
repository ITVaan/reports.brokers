import os
from shutil import copyfile
from datetime import datetime
import mysql.connector as mariadb
from openpyxl import load_workbook
from reports.brokers.api.selections import report1


sql = report1

data = []

conn = mariadb.connect(host='127.0.0.1', user='root', password='root', database='reports_data', charset='utf8')
cursor = conn.cursor(buffered=True)

res = cursor.execute(sql)
for broker_name, suppliers_count in cursor:
    data.append([broker_name, suppliers_count])

cursor.close()
conn.close()

templates_dir = 'templates'
result_dir = 'reports'
template_file_name = 'one.xlsx'

t = os.path.splitext(template_file_name)
result_file = os.path.join(result_dir, t[0] + '-' + datetime.now().strftime('%Y-%m-%d-%H-%M-%S') + t[1])

copyfile(os.path.join(templates_dir, template_file_name), result_file)

wb = load_workbook(filename=result_file)
ws = wb.active

row = 2

for (broker_name, suppliers_count) in data:
    ws.cell(row=row, column=1, value=broker_name)
    ws.cell(row=row, column=2, value=suppliers_count)
    row += 1
    print("{} - {}".format(broker_name, suppliers_count))

wb.save(result_file)
