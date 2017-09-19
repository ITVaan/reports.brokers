# coding=utf-8
from __future__ import division

import hashlib
from datetime import datetime
from logging import getLogger
from shutil import copyfile
from uuid import uuid4

import mysql.connector as mariadb
import os
from openpyxl import load_workbook

from reports.brokers.api.selections import *
from reports.brokers.utils import get_root_pwd

LOGGER = getLogger("{}.init".format(__name__))


class GeneratorOfReports(object):
    def __init__(self, start_report_period, end_report_period, report_number, user_name, password, config):
        # Report period
        self.start_report_period = datetime.strptime(str(start_report_period), '%d.%m.%Y')
        self.end_report_period = datetime.strptime(str(end_report_period), '%d.%m.%Y')

        # Authentication check
        self.report_number = report_number
        self.user_name = user_name
        self.password = hashlib.sha256(password).hexdigest()
        self.config = config

        self.data = []

        # Excel directories
        self.templates_dir = self.config_get("templates_dir")
        self.result_dir = self.config_get("result_dir")

        # DataBase connection
        self.conn = mariadb.connect(host=self.config_get("db_host"), user=self.config_get("db_user"),
                                    password=get_root_pwd(), database=self.config_get("database"),
                                    charset=self.config_get("db_charset") or 'utf8')
        self.cursor = self.conn.cursor(buffered=True, named_tuple=True)

        # Launching of reports generator
        self.user_id = self.auth()
        if self.user_id:
            self.date = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
            self.uuid = uuid4().hex
            self.start_reporting()
            self.logging()
            self.conn.commit()
            self.cursor.close()
            self.conn.close()
            self.wb.save(self.result_file)
        else:
            LOGGER.info('Permission denied')

    def config_get(self, name):
        return self.config.get(name)

    def auth(self):
        self.cursor.execute(auth, {'user_name': self.user_name, 'password': self.password})
        for i in self.cursor:
            if isinstance(i[0], int):
                return i[0]

    def start_reporting(self):
        # Report file creation
        template_file_name = '{}.xlsx'.format(self.report_number)
        file_format = os.path.splitext(template_file_name)[1]
        self.result_file = os.path.join(self.result_dir, self.filename(file_format))
        copyfile(os.path.join(self.templates_dir, template_file_name), self.result_file)
        self.wb = load_workbook(filename=self.result_file)
        self.ws = self.wb.active

        # Start
        LOGGER.info("Start reporting: rep_number={}".format(self.report_number))
        if self.report_number == '1':
            self.cursor.execute(report1, {'start_date': self.start_report_period, 'end_date': self.end_report_period})
            self.report_1()
        elif self.report_number == '2':
            self.cursor.execute(report2, {'start_date': self.start_report_period, 'end_date': self.end_report_period})
            self.report_2()
        elif self.report_number == '3':
            self.cursor.execute(report3, {'start_date': self.start_report_period, 'end_date': self.end_report_period})
            self.report_3()

    def logging(self):
        self.cursor.execute(logging, {'user_id': self.user_id,
                                      'report_type_id': self.report_number,
                                      'start_report_period': self.start_report_period,
                                      'end_report_period': self.end_report_period})

    def filename(self, file_format):
        return "{date}_report_number_{num}_{uid}_{uuid4}{ext}".format(date=self.date,
                                                                      num=str(self.report_number),
                                                                      uid=str(self.user_id),
                                                                      uuid4=self.uuid, ext=file_format)

    def report_1(self):
        for row in self.cursor:
            self.ws.append(row)

    def report_2(self):
        for row in self.cursor:
            self.ws.append(row)

    def report_3(self):
        self.data = [row for row in self.cursor]
        br_names = sorted(list(set(row.broker_name for row in self.data)))
        zero_quartiles = []
        fourth_quartiles = []

        for br_name in br_names:
            tmp = [row.bids_count for row in self.data if row.broker_name == br_name]
            zero_quartiles.append((br_name, min(tmp)))
            fourth_quartiles.append((br_name, max(tmp)))
        initial_row = 2

        for br_name in br_names:
            zero_q = filter(lambda x: x[0] == br_name, zero_quartiles)[0][1]
            fourth_q = filter(lambda x: x[0] == br_name, fourth_quartiles)[0][1]
            first_q = (fourth_q - zero_q) * 0.25
            second_q = (fourth_q - zero_q) * 0.5
            third_q = (fourth_q - zero_q) * 0.75

            zero_q_p = 0
            first_q_p = 0
            second_q_p = 0
            third_q_p = 0
            fourth_q_p = 0

            tmp = []
            tmp_tend = []
            for row in self.data:
                if row.broker_name == br_name:
                    tmp.append(row.bids_count)
                    tmp_tend.append(row.tenderers_identifier)
                    if zero_q == row.bids_count:
                        zero_q_p += 1
                    elif zero_q < row.bids_count <= first_q:
                        first_q_p += 1
                    elif first_q < row.bids_count <= second_q:
                        second_q_p += 1
                    elif second_q < row.bids_count <= third_q:
                        third_q_p += 1
                    elif third_q < row.bids_count <= fourth_q:
                        fourth_q_p += 1
            sum_bids_count = sum(tmp)
            num_tenderers = len(set(tmp_tend))
            average_part = "%.2f" % (sum_bids_count / num_tenderers)
            self.ws.append((br_name, average_part, 'quartiles', zero_q, first_q, second_q, third_q, fourth_q))
            self.ws.append((br_name, average_part, 'participants', zero_q, first_q, second_q, third_q, fourth_q))
            self.ws.merge_cells(start_row=initial_row, start_column=1, end_row=initial_row + 1, end_column=1)
            self.ws.merge_cells(start_row=initial_row, start_column=2, end_row=initial_row + 1, end_column=2)
            initial_row += 2
