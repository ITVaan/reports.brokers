# -*- coding: utf-8 -*-
from openpyxl import Workbook, load_workbook
from os import path
from shutil import copyfile
from gevent import sleep as gsleep


def copy_xls_file_from_template():
    templates_dir = 'reports/brokers/api/views/templates'
    result_dir = 'reports/brokers/tests/test_reports'
    template_file_name = '1.xlsx'
    file_format = path.splitext(template_file_name)
    result_file = path.join(result_dir, '2017-08-05-17-00-00' + '_report-number=' +
                            file_format[0] + file_format[1])
    copyfile(path.join(templates_dir, template_file_name), result_file)
    return result_file


def create_example_worksheet():
    wb_expected = Workbook()
    ws_expected = wb_expected.active
    ws_expected.title = u"Лист1"
    ws_expected.cell(row=1, column=1, value="Площадка")
    ws_expected.cell(row=1, column=2, value="Кількість нових учасників")
    ws_expected.cell(row=2, column=1, value="prom.ua")
    ws_expected.cell(row=2, column=2, value=1)
    return ws_expected


def load_and_fill_result_workbook(data, result_file):
    res = load_workbook(filename=result_file)
    ws = res.active
    row = 2
    for (broker_name, suppliers_count) in data:
        ws.cell(row=row, column=1, value=broker_name)
        ws.cell(row=row, column=2, value=suppliers_count)
        row += 1
    return res


def custom_sleep(seconds=0):
    return gsleep(seconds=0)
