# -*- coding: utf-8 -*-
from datetime import datetime, timedelta

import hypothesis.strategies as st
from hypothesis import assume, given
from openpyxl import load_workbook

from reports.brokers.api.selections import report1
from reports.brokers.tests.base_db_test import BaseDbTestCase
from reports.brokers.tests.utils import copy_xls_file_from_template, create_example_worksheet


class TestDataBaseConnection(BaseDbTestCase):
    def test_create_empty_xls(self):
        cursor = self.conn.cursor(buffered=True, named_tuple=True)
        result_file = copy_xls_file_from_template()
        wb = load_workbook(filename=result_file)
        ws = wb.active
        cursor.execute(report1, {'start_date': '01.05.2017', 'end_date': '01.06.2017'})
        cursor.close()
        for row in cursor:
            ws.append(row)

    @given(st.datetimes(min_value=datetime(2000, 1, 1, 0, 0)), st.datetimes(min_value=datetime(2000, 1, 1, 0, 0)))
    def test_create_nonempty_xls(self, start, end):
        assume(end > start + timedelta(seconds=1))
        modified = start
        cursor = self.conn.cursor(buffered=True, named_tuple=True)
        cursor.execute("""
            insert into `tenders` (`original_id`,`status_id`,`broker_id`,`date_modified`,`enquiry_start_date`,
              `enquiry_end_date`) 
            values ('111', 1, 1, '{}', '{}', '{}')  
        """.format(modified, start, end))
        tender_db_id = cursor.lastrowid
        cursor.execute("INSERT INTO `tenderers` (`identifier`, `scheme`) VALUES ('12345', '1234');")
        tenderer_db_id = cursor.lastrowid
        cursor.execute(
            "INSERT INTO `bids` (`original_id`, `tender_id`, `status_id`) VALUES ('111', {}, 1)".format(tender_db_id))
        bid_db_id = cursor.lastrowid
        cursor.execute(
            "INSERT INTO `tenderers_bids` (`tenderer_id`, `bid_id`) VALUES ({}, {})".format(tenderer_db_id, bid_db_id))
        cursor.execute(report1, {'start_date': start, 'end_date': end})
        result_file = copy_xls_file_from_template()
        wb = load_workbook(filename=result_file)
        ws = wb.active
        for row in cursor:
            ws.append(row)
        cursor.close()
        ws_expected = create_example_worksheet()
        self.assertEqual([row.value for col in ws for row in col], [row.value for col in ws_expected for row in col])
