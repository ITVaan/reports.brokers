# coding=utf-8
from ConfigParser import SafeConfigParser
from datetime import datetime
from shutil import copyfile
from unittest import TestCase

import os
import re

import mysql.connector as mariadb
from openpyxl import load_workbook


def execute_scripts_from_file(cursor, filename):
    print("Execution in process")
    # Open and read the file as a single buffer
    fd = open(filename, 'r')
    sql_file = fd.read()
    fd.close()
    # Find special delimiters
    delimiters = re.compile('DELIMITER *(\S*)', re.I)
    result = delimiters.split(sql_file)
    # Insert default delimiter and separate delimiters and sql
    result.insert(0, ';')
    delimiter = result[0::2]
    section = result[1::2]
    for i in range(len(delimiter)):
        queries = section[i].split(delimiter[i])
        for query in queries:
            if not query.strip():
                continue
            cursor.execute(query)


sql = """SELECT grp1.first_broker, COUNT(grp1.identifier) AS new_tenderers_count
FROM
    (
        SELECT
            ts.`identifier`,
            (
                SELECT br2.code
                FROM
                    tenders t2
                    LEFT JOIN `brokers` br2 ON br2.`id` = t2.`broker_id`
                    LEFT JOIN bids b2 ON b2.`tender_id` = t2.id
                    LEFT JOIN `tenderers_bids` tb2 ON tb2.`bid_id` = b2.id
                WHERE tb2.tenderer_id = ts.`id`
                ORDER BY b2.bid_date
                LIMIT 1) AS first_broker,
                COUNT(t.id) AS tenders_count
        FROM
            tenders t
            LEFT JOIN `brokers` br ON br.`id` = t.`broker_id`
            LEFT JOIN bids b ON b.`tender_id` = t.id
            LEFT JOIN `tenderers_bids` tb ON tb.`bid_id` = b.id
            LEFT JOIN tenderers ts ON ts.`id` = tb.`tenderer_id`
        WHERE ts.`id` IS NOT NULL
        GROUP BY ts.`identifier`
        HAVING COUNT(t.id) > 1
        ORDER BY ts.`id`
    ) AS grp1
GROUP BY grp1.first_broker
ORDER BY 2 DESC""";


def get_item_from_section(conf_parser, section, item):
    """
    :type conf_parser: SafeConfigParser
    :param conf_parser: Config parser with file already read into it
    :type section: str
    :param section: String, name of section
    :type item: str
    :param item: String, name of option
    :type return: str
    :return: item or empty string if no such section
    """
    if conf_parser.has_option(section, item):
        return conf_parser.get(section, item)
    else:
        return ''


class TestDataBaseConnection(TestCase):
    @classmethod
    def setUpClass(cls):
        conf_parser = SafeConfigParser()
        conf_parser.read("auth.ini")
        password = get_item_from_section(conf_parser, "Database", "root_password")
        cls.conn = mariadb.connect(host='127.0.0.1', user='root', password=password)

    @classmethod
    def tearDownClass(cls):
        cls.conn.close()

    def setUp(self):
        cursor = self.conn.cursor(buffered=True)
        cursor.execute("""DROP DATABASE IF EXISTS reports_data_test;""")
        cursor.execute("""CREATE DATABASE IF NOT EXISTS reports_data_test;""")
        cursor.execute("""USE reports_data_test""")
        execute_scripts_from_file(cursor, "reports/brokers/database/reports_data_dev.sql")
        cursor.close()

    def tearDown(self):
        cursor = self.conn.cursor(buffered=True)
        cursor.execute("""DROP DATABASE IF EXISTS reports_data_test;""")
        cursor.close()

    def test_create_xls(self):
        cursor = self.conn.cursor(buffered=True)
        res = cursor.execute(sql)
        data = []
        for broker_name, suppliers_count in cursor:
            data.append([broker_name, suppliers_count])
        cursor.close()
        templates_dir = 'reports/templates'
        result_dir = 'reports'
        template_file_name = 'one.xlsx'
        t = os.path.splitext(template_file_name)
        result_file = os.path.join(result_dir, t[0] + '-' + datetime.now().strftime('%Y-%m-%d-%H-%M-%S') + t[1])
        copyfile(os.path.join(templates_dir, template_file_name), result_file)
        wb = load_workbook(filename=result_file)
        ws = wb.active
        row = 2
        for (broker_name, suppliers_count) in data:
            ws.cell(row=row, column=1, value=broker_name)
            ws.cell(row=row, column=2, value=suppliers_count)
            row += 1
            print("{} - {}".format(broker_name, suppliers_count))
        wb.save(result_file)
