# -*- coding: utf-8 -*-
from ConfigParser import SafeConfigParser
from datetime import datetime
from shutil import copyfile
from uuid import uuid4

import os
from re import compile, I

from gevent import sleep as gsleep
from openpyxl import Workbook, load_workbook

test_config = {
    'db_user': 'root',
    'db_password': 'root',
    'db_host': 'localhost',
    'database': 'reports_data_test',
    'db_charset': 'utf8',
    'templates_dir': "reports/brokers/api/views/templates",
    'result_dir': os.path.join(os.path.dirname(os.path.realpath(__file__)), "test_reports")
}

config = SafeConfigParser()
config.read(os.path.join(os.path.dirname(__file__), "config.ini"))


def config_get(name):
    return config.get('app:api', name)


def filename(report_number, user_id, file_format):
    return "{date}_start_date_{start_date}_end_date_{end_date}_report_number_{num}_{uid}_{uuid4}{ext}".format(
        date=datetime.now().strftime('%Y-%m-%d-%H-%M-%S'),
        start_date='2016-05-02 13:00:00',
        end_date='2017-05-02 13:00:00',
        num=str(report_number),
        uid=str(user_id),
        uuid4=uuid4().hex, ext=file_format)


def copy_xls_file_from_template():
    template_file_name = '1.xlsx'
    result_file = os.path.join(test_config.get('result_dir'), filename(1, 1, ".xlsx"))
    copyfile(os.path.join(test_config.get('templates_dir'), template_file_name), result_file)
    return result_file


def test_purge():
    for file in os.listdir(test_config.get('result_dir')):
        if ".xls" in file or ".xlsx" in file:
            os.remove(os.path.join(test_config.get('result_dir'), file))


def create_example_worksheet():
    wb_expected = Workbook()
    ws_expected = wb_expected.active
    ws_expected.title = u"Лист1"
    ws_expected.cell(row=1, column=1, value="Broker")
    ws_expected.cell(row=1, column=2, value="Number of new suppliers")
    ws_expected.cell(row=2, column=1, value="prom.ua")
    ws_expected.cell(row=2, column=2, value=1)
    return ws_expected


def load_and_fill_result_workbook(data, result_file):
    res = load_workbook(filename=result_file)
    ws = res.active
    row = 2
    for (broker_name, suppliers_count) in data:
        ws.cell(row=row, column=1, value=broker_name)
        ws.cell(row=row, column=2, value=suppliers_count)
        row += 1
    return res


def custom_sleep(seconds=0):
    return gsleep(seconds=0)


def execute_scripts_from_file(cursor, filename):
    # Open and read the file as a single buffer
    fd = open(filename, 'r')
    sql_file = fd.read()
    fd.close()
    # Find special delimiters
    delimiters = compile('DELIMITER *(\S*)', I)
    result = delimiters.split(sql_file)
    # Insert default delimiter and separate delimiters and sql
    result.insert(0, ';')
    delimiter = result[0::2]
    section = result[1::2]
    for i in range(len(delimiter)):
        queries = section[i].split(delimiter[i])
        for query in queries:
            if not query.strip():
                continue
            cursor.execute(query)
